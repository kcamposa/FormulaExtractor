import openpyxl

results = []
formulas = []

def ReadFile():
    theFile = openpyxl.load_workbook( 'multiply.xlsx' )
    allSheetNames = theFile.sheetnames

    print( "===========================================" )
    print( "Extracting data from {}".format( theFile.sheetnames ) )
    print( "===========================================" )

    for sheet in allSheetNames:
        currentSheet = theFile[sheet]
        for row in range( 1, currentSheet.max_row + 1 ):
            for column in "A": 
                cell_name = "{}{}".format( column, row )       
                operation = currentSheet[cell_name].value
                saveFormula(operation)

def saveFormula(ope):
    if ope != 'FIN':
        if ope == None:
            ope = '***'
            formulas.append(ope)
        else:
            formulas.append(ope)

def ExtractResults(): # 0*48+2.0
    cont = 0
    for n in formulas:
        if n != "FIN":
            if n != "***":
                left = 0
                core = 48
                right = 0

                data = n.index('Y')

                left = int(n[:data])
                right = float(n[data+1:])

                results.append( left * core + right )
                cont = cont + 1
            else:
                results.append("***")

def WriteResults():
    for n in results:
        file = open('results.txt', 'a')
        file.write( str(n) + "\n")
        file.close()


ReadFile()
ExtractResults()
WriteResults()
