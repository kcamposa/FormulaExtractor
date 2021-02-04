import openpyxl
import array

cells = []
newFormula = []

def ReadFile():
    theFile = openpyxl.load_workbook( 'columnA.xlsx' )
    allSheetNames = theFile.sheetnames

    print( "===========================================" )
    print( "Extracting data from {}".format( theFile.sheetnames ) )
    print( "===========================================" )

    for sheet in allSheetNames:
        currentSheet = theFile[sheet]
        for row in range( 1, currentSheet.max_row + 1 ):
            for column in "A":  
                cell_name = "{}{}".format( column, row )       
                cell = currentSheet[cell_name].value
                captureFormula( cell )

def captureFormula(cell):
    if cell != "":
        cells.append(cell)

def splitFormula():
    for n in cells:
        cellName = n.replace(u'\xa0', u' ')
        x = cellName.replace('\n', ',')
        ultraSplit = " ".join(x.split())
        newX = ultraSplit.replace(' ', ',')
        newFormula.append(newX)

def WriteFormulaTXT():
    cont = 0
    for n in newFormula:
        file = open('formulas.txt', 'a')
        file.write(str(cont) + ","+ n + "\n")
        file.close()
        cont = cont + 1      

ReadFile()
splitFormula()
WriteFormulaTXT()
